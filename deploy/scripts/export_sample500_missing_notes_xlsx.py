#!/usr/bin/env python3
"""Export Excel: sample-500 missing DailyNote@DOS + orphan DOS + lag/wrong-case audit."""
from __future__ import annotations

import argparse
import csv
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

_DATE_IN_NAME = re.compile(r"(20\d{2}-\d{2}-\d{2})")
_CLINICAL_TYPES = frozenset(
    {"daily_note", "poc", "discharge", "progress", "eval", "other_clinical"}
)


def _load_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as fh:
        return list(csv.DictReader(fh))


def _is_daily_note(row: dict[str, str]) -> bool:
    did = (row.get("daily_note_id") or "").strip()
    nfile = (row.get("note_file") or "").lower()
    return did.startswith("DN") or "dailynote" in nfile


def _parse_date(value: str) -> datetime | None:
    raw = (value or "")[:10]
    if not raw:
        return None
    try:
        return datetime.fromisoformat(raw)
    except ValueError:
        return None


def _case_has_dn_pdf(cases_root: Path, facility_id: str, case_id: str) -> tuple[bool, int]:
    d = cases_root / facility_id / case_id / "daily_notes"
    if not d.is_dir():
        return False, 0
    pdfs = list(d.glob("*.pdf"))
    dn = [p for p in pdfs if "dailynote" in p.name.lower() or "DN" in p.name.upper()]
    count = len(dn) if dn else len(pdfs)
    return (count > 0), count


def _classify_pdf_name(name: str) -> str | None:
    """Return clinical artifact type from WebPT PDF filename, or None."""
    n = name.lower()
    if "dailynote" in n or n.startswith("dn") or "-dn" in n or "_dn" in n:
        return "daily_note"
    if "discharge" in n or "-dq" in n or "_dq" in n:
        return "discharge"
    if "plan" in n and "care" in n:
        return "poc"
    if "-po" in n or "_po" in n:
        if "progress" not in n and "dailynote" not in n:
            return "poc"
    if "progress" in n:
        return "progress"
    if "eval" in n or "-ie" in n or "_ie" in n:
        return "eval"
    if "-wo" in n or "_wo" in n or "addendum" in n:
        return "other_clinical"
    return None


def _scan_case_clinical(
    cases_root: Path, facility_id: str, case_id: str, visit_dos: datetime | None
) -> dict[str, object]:
    """Scan case PDFs for POC/Discharge/clinical artifacts vs visit DOS."""
    empty = {
        "has_poc_on_case": "N",
        "nearest_poc_dos": "",
        "poc_delta_days": "",
        "has_discharge_on_case": "N",
        "nearest_discharge_dos": "",
        "discharge_delta_days": "",
        "any_clinical_exact_dos": "N",
        "clinical_exact_dos_types": "",
        "has_poc_or_discharge_exact_dos": "N",
    }
    case_dir = cases_root / facility_id / case_id
    if not case_dir.is_dir():
        return empty

    dates_by_type: dict[str, list[datetime]] = defaultdict(list)
    exact_types: set[str] = set()
    case_types: set[str] = set()

    for pdf in case_dir.rglob("*.pdf"):
        kind = _classify_pdf_name(pdf.name)
        if kind is None:
            continue
        case_types.add(kind)
        for ds in _DATE_IN_NAME.findall(pdf.name):
            dt = _parse_date(ds)
            if dt is None:
                continue
            dates_by_type[kind].append(dt)
            if visit_dos is not None and dt.date() == visit_dos.date():
                exact_types.add(kind)

    def nearest_for(kind: str) -> tuple[datetime | None, int | None]:
        if visit_dos is None:
            return None, None
        dates = dates_by_type.get(kind, [])
        if not dates:
            return None, None
        best = min(dates, key=lambda d: (abs((d - visit_dos).days), (d - visit_dos).days))
        return best, (best - visit_dos).days

    poc_dt, poc_delta = nearest_for("poc")
    dq_dt, dq_delta = nearest_for("discharge")
    clinical_exact = sorted(exact_types & _CLINICAL_TYPES)

    return {
        "has_poc_on_case": "Y" if "poc" in case_types else "N",
        "nearest_poc_dos": poc_dt.date().isoformat() if poc_dt else "",
        "poc_delta_days": poc_delta if poc_delta is not None else "",
        "has_discharge_on_case": "Y" if "discharge" in case_types else "N",
        "nearest_discharge_dos": dq_dt.date().isoformat() if dq_dt else "",
        "discharge_delta_days": dq_delta if dq_delta is not None else "",
        "any_clinical_exact_dos": "Y" if clinical_exact else "N",
        "clinical_exact_dos_types": ",".join(clinical_exact),
        "has_poc_or_discharge_exact_dos": (
            "Y" if exact_types & {"poc", "discharge"} else "N"
        ),
    }


def _load_sf_billing_index(path: Path) -> dict[tuple[str, str], str]:
    """Map (EMR_ID, DATE_OF_SERVICE) -> STATUS from Snowflake billing pull."""
    if not path.is_file():
        return {}
    out: dict[tuple[str, str], str] = {}
    with path.open(encoding="utf-8-sig", newline="") as fh:
        for row in csv.DictReader(fh):
            emr = (row.get("EMR_ID") or "").strip()
            dos = (row.get("DATE_OF_SERVICE") or "")[:10]
            if not emr or not dos:
                continue
            status = (row.get("STATUS") or "").strip()
            key = (emr, dos)
            # Prefer Paid if duplicates
            prev = out.get(key)
            if prev is None or (status == "Paid" and prev != "Paid"):
                out[key] = status
    return out


def _nearest_dn(
    visit_dos: datetime, note_dates: list[datetime]
) -> tuple[datetime | None, int | None]:
    """Prefer smallest non-negative lag; else closest overall."""
    if not note_dates:
        return None, None
    best: datetime | None = None
    best_delta: int | None = None
    for nd in note_dates:
        delta = (nd - visit_dos).days
        if best is None:
            best, best_delta = nd, delta
            continue
        assert best_delta is not None
        # Prefer non-negative closer; else closer absolute
        if delta >= 0 and (best_delta < 0 or delta < best_delta):
            best, best_delta = nd, delta
        elif best_delta < 0 and delta < 0 and delta > best_delta:
            best, best_delta = nd, delta
        elif best_delta < 0 and delta >= 0:
            best, best_delta = nd, delta
        elif abs(delta) < abs(best_delta) and not (best_delta >= 0 and delta < 0):
            if best_delta < 0 or delta >= 0:
                best, best_delta = nd, delta
    return best, best_delta


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument(
        "--sample-csv",
        type=Path,
        default=Path("/home/abdu/sf_eval/checked_out_gap_sample_500.csv"),
    )
    ap.add_argument(
        "--notes-csv",
        type=Path,
        default=Path(
            "/data/exports/side_by_side_case/sample500_extracted/daily_notes.csv"
        ),
    )
    ap.add_argument(
        "--schedule-csv",
        type=Path,
        default=Path("/data/exports/side_by_side_case/schedule/schedule_cases.csv"),
    )
    ap.add_argument(
        "--cases-root",
        type=Path,
        default=Path("/data/exports/side_by_side_case/cases"),
    )
    ap.add_argument(
        "--out",
        type=Path,
        default=Path("/home/abdu/sf_eval/checked_out_gap_sample_500_missing_notes.xlsx"),
    )
    ap.add_argument(
        "--billing-csv",
        type=Path,
        default=Path("/home/abdu/sf_eval/billing_2026-01-01_to_now.csv"),
        help="Snowflake ALL_BILLING_DATA pull (EMR_ID + DATE_OF_SERVICE + STATUS)",
    )
    args = ap.parse_args()

    sample = _load_csv(args.sample_csv)
    notes = _load_csv(args.notes_csv)
    schedule = _load_csv(args.schedule_csv)
    sf_index = _load_sf_billing_index(args.billing_csv)

    sched_dos_by_case: dict[str, set[str]] = defaultdict(set)
    cases_by_pid: dict[str, set[str]] = defaultdict(set)
    for row in schedule:
        case_id = (row.get("case_id") or "").strip()
        pid = (row.get("patient_id") or "").strip()
        dos = (row.get("dos") or "")[:10]
        if case_id and dos:
            sched_dos_by_case[case_id].add(dos)
        if pid and case_id:
            cases_by_pid[pid].add(case_id)

    dn_by_case: dict[str, list[dict[str, str]]] = defaultdict(list)
    dn_by_pid: dict[str, list[dict[str, str]]] = defaultdict(list)
    dn_for_visit: set[tuple[str, str, str]] = set()
    for row in notes:
        if not _is_daily_note(row):
            continue
        case_id = (row.get("case_id") or "").strip()
        pid = (row.get("patient_id") or "").strip()
        dos = (row.get("date_of_daily_note") or "")[:10]
        if not case_id:
            continue
        dn_by_case[case_id].append(row)
        if pid:
            dn_by_pid[pid].append(row)
        if pid and dos:
            dn_for_visit.add((case_id, pid, dos))

    sample_gap_dos_by_case: dict[str, list[str]] = defaultdict(list)

    missing_rows: list[dict[str, object]] = []
    has_dn = 0
    lag_counts = Counter()  # 1,2,3 among missing
    wrong_case_exact = 0
    wrong_case_lag13 = 0
    multi_case_missing = 0
    no_dn_on_case = 0
    clinical_counts = Counter()

    for r in sample:
        case_id = (r.get("case_id") or "").strip()
        pid = (r.get("patient_id") or r.get("emr_id") or "").strip()
        dos = (r.get("dos") or "")[:10]
        fac = (r.get("facility_id") or "").strip()
        key = (case_id, pid, dos)
        if key in dn_for_visit:
            has_dn += 1
            continue

        sample_gap_dos_by_case[case_id].append(dos)
        has_pdf, pdf_count = _case_has_dn_pdf(args.cases_root, fac, case_id)
        extract_dn_count = len(dn_by_case.get(case_id, []))
        case_dn_count = max(extract_dn_count, pdf_count)
        if case_dn_count == 0:
            no_dn_on_case += 1

        sched = sched_dos_by_case.get(case_id, set())
        orphan_count = 0
        for n in dn_by_case.get(case_id, []):
            ndos = (n.get("date_of_daily_note") or "")[:10]
            if ndos and ndos not in sched:
                orphan_count += 1

        visit_dt = _parse_date(dos)
        same_case_dates = [
            d
            for n in dn_by_case.get(case_id, [])
            if (d := _parse_date(n.get("date_of_daily_note") or ""))
        ]
        nearest, delta = (
            _nearest_dn(visit_dt, same_case_dates) if visit_dt else (None, None)
        )
        deltas_same = [
            (d - visit_dt).days for d in same_case_dates if visit_dt is not None
        ]
        has_lag_1 = "Y" if 1 in deltas_same else "N"
        has_lag_2 = "Y" if 2 in deltas_same else "N"
        has_lag_3 = "Y" if 3 in deltas_same else "N"
        has_lag_13 = "Y" if any(x in (1, 2, 3) for x in deltas_same) else "N"
        if has_lag_13 == "Y":
            # count smallest positive among 1..3
            pos = min(x for x in deltas_same if x in (1, 2, 3))
            lag_counts[pos] += 1

        # Other case same patient
        other_exact = False
        other_lag13 = False
        for n in dn_by_pid.get(pid, []):
            c2 = (n.get("case_id") or "").strip()
            if c2 == case_id:
                continue
            ndos = (n.get("date_of_daily_note") or "")[:10]
            if ndos == dos:
                other_exact = True
            nd = _parse_date(ndos)
            if visit_dt and nd and (nd - visit_dt).days in (1, 2, 3):
                other_lag13 = True
        if other_exact:
            wrong_case_exact += 1
        if other_lag13:
            wrong_case_lag13 += 1

        patient_case_count = len(cases_by_pid.get(pid, set()))
        if patient_case_count > 1:
            multi_case_missing += 1

        clinical = _scan_case_clinical(args.cases_root, fac, case_id, visit_dt)
        if clinical["has_poc_on_case"] == "Y":
            clinical_counts["has_poc_on_case"] += 1
        if clinical["has_discharge_on_case"] == "Y":
            clinical_counts["has_discharge_on_case"] += 1
        if clinical["any_clinical_exact_dos"] == "Y":
            clinical_counts["any_clinical_exact_dos"] += 1
        if clinical["has_poc_or_discharge_exact_dos"] == "Y":
            clinical_counts["has_poc_or_discharge_exact_dos"] += 1
        poc_d = clinical["poc_delta_days"]
        dq_d = clinical["discharge_delta_days"]
        if clinical["has_poc_on_case"] == "Y" or clinical["has_discharge_on_case"] == "Y":
            deltas_poc_dq = [
                abs(int(x))
                for x in (poc_d, dq_d)
                if x != "" and x is not None
            ]
            if deltas_poc_dq:
                nearest_abs = min(deltas_poc_dq)
                if nearest_abs <= 3:
                    clinical_counts["nearest_poc_or_dq_abs_1_to_3"] += 1
                if nearest_abs <= 30:
                    clinical_counts["nearest_poc_or_dq_abs_le_30"] += 1
            else:
                clinical_counts["poc_or_dq_undated"] += 1
        else:
            clinical_counts["no_poc_or_discharge_on_case"] += 1

        emr = (r.get("emr_id") or pid).strip()
        sf_status = sf_index.get((emr, dos), "")
        in_sf = "Y" if (emr, dos) in sf_index else "N"
        if in_sf == "Y":
            clinical_counts["in_sf_billing"] += 1
            if sf_status == "Paid":
                clinical_counts["sf_status_paid"] += 1

        missing_rows.append(
            {
                "patient_id": pid,
                "emr_id": emr,
                "patient_name": (r.get("patient") or "").strip(),
                "facility_id": fac,
                "facility_name": (r.get("facility_name") or "").strip(),
                "case_id": case_id,
                "dos": dos,
                "unit_id": (r.get("unit_id") or "").strip(),
                "sf_check": (r.get("sf_check") or "").strip(),
                "visit_status": (r.get("visit_status") or "").strip(),
                "case_has_any_daily_note_pdf": "Y" if has_pdf or extract_dn_count else "N",
                "case_daily_note_count": case_dn_count,
                "case_notes_with_dos_not_in_schedule_count": orphan_count,
                "nearest_dn_dos_same_case": nearest.date().isoformat() if nearest else "",
                "days_note_minus_visit": delta if delta is not None else "",
                "has_dn_lag_1": has_lag_1,
                "has_dn_lag_2": has_lag_2,
                "has_dn_lag_3": has_lag_3,
                "has_dn_lag_1_to_3": has_lag_13,
                "patient_case_count_in_schedule": patient_case_count,
                "has_dn_exact_on_other_case": "Y" if other_exact else "N",
                "has_dn_lag_1_to_3_on_other_case": "Y" if other_lag13 else "N",
                "has_poc_on_case": clinical["has_poc_on_case"],
                "nearest_poc_dos": clinical["nearest_poc_dos"],
                "poc_delta_days": clinical["poc_delta_days"],
                "has_discharge_on_case": clinical["has_discharge_on_case"],
                "nearest_discharge_dos": clinical["nearest_discharge_dos"],
                "discharge_delta_days": clinical["discharge_delta_days"],
                "any_clinical_exact_dos": clinical["any_clinical_exact_dos"],
                "clinical_exact_dos_types": clinical["clinical_exact_dos_types"],
                "has_poc_or_discharge_exact_dos": clinical[
                    "has_poc_or_discharge_exact_dos"
                ],
                "in_sf_billing": in_sf,
                "sf_status": sf_status,
            }
        )

    # Sheet 2 orphan notes
    sample_case_ids = {(r.get("case_id") or "").strip() for r in sample}
    orphan_rows: list[dict[str, object]] = []
    cases_with_orphan: set[str] = set()
    for case_id in sorted(sample_case_ids):
        if not case_id:
            continue
        sched = sched_dos_by_case.get(case_id, set())
        gap_list = sorted(set(sample_gap_dos_by_case.get(case_id, [])))
        gap_str = ",".join(gap_list)
        for n in dn_by_case.get(case_id, []):
            ndos = (n.get("date_of_daily_note") or "")[:10]
            if not ndos or ndos in sched:
                continue
            cases_with_orphan.add(case_id)
            orphan_rows.append(
                {
                    "patient_id": (n.get("patient_id") or "").strip(),
                    "case_id": case_id,
                    "facility_id": (n.get("facility_id") or "").strip(),
                    "note_dos": ndos,
                    "daily_note_id": (n.get("daily_note_id") or "").strip(),
                    "note_file": (n.get("note_file") or "").strip(),
                    "sample_gap_dos_list": gap_str,
                }
            )

    n_miss = len(missing_rows)
    lag13_total = sum(lag_counts[i] for i in (1, 2, 3))

    wb = Workbook()

    # Sheet 1 — missing + audit cols
    ws1 = wb.active
    ws1.title = "missing_daily_note_for_dos"
    headers1 = [
        "patient_id",
        "emr_id",
        "patient_name",
        "facility_id",
        "facility_name",
        "case_id",
        "dos",
        "unit_id",
        "sf_check",
        "visit_status",
        "case_has_any_daily_note_pdf",
        "case_daily_note_count",
        "case_notes_with_dos_not_in_schedule_count",
        "nearest_dn_dos_same_case",
        "days_note_minus_visit",
        "has_dn_lag_1",
        "has_dn_lag_2",
        "has_dn_lag_3",
        "has_dn_lag_1_to_3",
        "patient_case_count_in_schedule",
        "has_dn_exact_on_other_case",
        "has_dn_lag_1_to_3_on_other_case",
        "has_poc_on_case",
        "nearest_poc_dos",
        "poc_delta_days",
        "has_discharge_on_case",
        "nearest_discharge_dos",
        "discharge_delta_days",
        "any_clinical_exact_dos",
        "clinical_exact_dos_types",
        "has_poc_or_discharge_exact_dos",
        "in_sf_billing",
        "sf_status",
    ]
    ws1.append(headers1)
    for cell in ws1[1]:
        cell.font = Font(bold=True)
    for row in missing_rows:
        ws1.append([row[h] for h in headers1])

    # Sheet 2
    ws2 = wb.create_sheet("case_notes_dos_not_in_schedule")
    headers2 = [
        "patient_id",
        "case_id",
        "facility_id",
        "note_dos",
        "daily_note_id",
        "note_file",
        "sample_gap_dos_list",
    ]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for row in orphan_rows:
        ws2.append([row[h] for h in headers2])

    # Sheet 3 summary
    ws3 = wb.create_sheet("summary")
    ws3.append(["metric", "value"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
    for k, v in [
        ("sample_n", len(sample)),
        ("missing_dn_for_dos", n_miss),
        ("has_dn_for_dos", has_dn),
        ("cases_with_orphan_note_dos", len(cases_with_orphan)),
        ("orphan_note_rows", len(orphan_rows)),
        ("definition_has_dn", "case_id+patient_id+DOS and (DN* id or DailyNote filename)"),
        (
            "definition_orphan",
            "DailyNote on sample case whose note DOS not in schedule_cases for that case_id",
        ),
        ("cohort", "checked_out_gap_sample_500"),
    ]:
        ws3.append([k, v])

    # Sheet 4 summary_causes
    ws4 = wb.create_sheet("summary_causes")
    ws4.append(["finding", "count", "pct_of_497_missing", "verdict"])
    for cell in ws4[1]:
        cell.font = Font(bold=True)

    def pct(n: int) -> str:
        return f"{round(100.0 * n / max(n_miss, 1), 1)}%"

    cause_rows = [
        ("Missing exact note_date == dos", n_miss, pct(n_miss), "baseline"),
        (
            "Case has some DailyNote (other DOS)",
            n_miss - no_dn_on_case,
            pct(n_miss - no_dn_on_case),
            "notes exist elsewhere on case",
        ),
        (
            "Case has no DailyNote at all",
            no_dn_on_case,
            pct(no_dn_on_case),
            "true empty case",
        ),
        ("DN same case DOS+1", lag_counts[1], pct(lag_counts[1]), "strict-date artifact"),
        ("DN same case DOS+2", lag_counts[2], pct(lag_counts[2]), "strict-date artifact"),
        ("DN same case DOS+3", lag_counts[3], pct(lag_counts[3]), "strict-date artifact"),
        (
            "DN same case any +1..+3",
            lag13_total,
            pct(lag13_total),
            "lag recovers only ~4%",
        ),
        (
            "DN other case same patient exact DOS",
            wrong_case_exact,
            pct(wrong_case_exact),
            "wrong_case REJECTED",
        ),
        (
            "DN other case same patient +1..+3",
            wrong_case_lag13,
            pct(wrong_case_lag13),
            "wrong_case REJECTED",
        ),
        (
            "Missing rows where patient has >1 schedule case",
            multi_case_missing,
            pct(multi_case_missing),
            "multi-case exists; does not explain via other-case DN",
        ),
        (
            "Dominant cause",
            n_miss - lag13_total,
            pct(n_miss - lag13_total),
            "no near-DOS DailyNote on mapped case",
        ),
    ]
    for row in cause_rows:
        ws4.append(list(row))

    ws4.append([])
    ws4.append(["verdict_wrong_case", "rejected"])
    ws4.append(["verdict_lag_1_to_3_recovers", f"{lag13_total}/{n_miss}"])
    ws4.append(
        [
            "verdict_dominant",
            "no near-DOS DailyNote on mapped case (~96% after lag window)",
        ]
    )

    # Sheet 5 lag_and_case_audit = same audit cols for convenience (full missing rows)
    ws5 = wb.create_sheet("lag_and_case_audit")
    headers5 = [
        "patient_id",
        "case_id",
        "dos",
        "nearest_dn_dos_same_case",
        "days_note_minus_visit",
        "has_dn_lag_1",
        "has_dn_lag_2",
        "has_dn_lag_3",
        "has_dn_lag_1_to_3",
        "patient_case_count_in_schedule",
        "has_dn_exact_on_other_case",
        "has_dn_lag_1_to_3_on_other_case",
        "case_daily_note_count",
    ]
    ws5.append(headers5)
    for cell in ws5[1]:
        cell.font = Font(bold=True)
    for row in missing_rows:
        ws5.append([row[h] for h in headers5])

    # Sheet 6 clinical artifacts + SF billing (report-only; TMG stays DailyNote-only)
    ws6 = wb.create_sheet("summary_clinical_artifacts")
    ws6.append(["finding", "count", "pct_of_missing", "note"])
    for cell in ws6[1]:
        cell.font = Font(bold=True)

    def pct_m(n: int) -> str:
        return f"{round(100.0 * n / max(n_miss, 1), 1)}%"

    clinical_summary = [
        (
            "Missing exact DailyNote note_date == dos",
            n_miss,
            pct_m(n_miss),
            "baseline sheet missing_daily_note_for_dos",
        ),
        (
            "Case has some Plan of Care PDF (PO*)",
            clinical_counts["has_poc_on_case"],
            pct_m(clinical_counts["has_poc_on_case"]),
            "case-level; date usually != visit DOS",
        ),
        (
            "Case has some Discharge PDF",
            clinical_counts["has_discharge_on_case"],
            pct_m(clinical_counts["has_discharge_on_case"]),
            "case-level",
        ),
        (
            "No POC or Discharge on case",
            clinical_counts["no_poc_or_discharge_on_case"],
            pct_m(clinical_counts["no_poc_or_discharge_on_case"]),
            "",
        ),
        (
            "Any clinical PDF filename date == visit DOS",
            clinical_counts["any_clinical_exact_dos"],
            pct_m(clinical_counts["any_clinical_exact_dos"]),
            "DN|POC|DQ|progress|eval|other_clinical",
        ),
        (
            "POC or Discharge exact DOS",
            clinical_counts["has_poc_or_discharge_exact_dos"],
            pct_m(clinical_counts["has_poc_or_discharge_exact_dos"]),
            "does not close DailyNote gap",
        ),
        (
            "Nearest POC/DQ within +/-1..3 days",
            clinical_counts["nearest_poc_or_dq_abs_1_to_3"],
            pct_m(clinical_counts["nearest_poc_or_dq_abs_1_to_3"]),
            "",
        ),
        (
            "Nearest POC/DQ within +/-30 days",
            clinical_counts["nearest_poc_or_dq_abs_le_30"],
            pct_m(clinical_counts["nearest_poc_or_dq_abs_le_30"]),
            "",
        ),
        (
            "In Snowflake ALL_BILLING_DATA (EMR+DOS)",
            clinical_counts["in_sf_billing"],
            pct_m(clinical_counts["in_sf_billing"]),
            f"billing_csv={args.billing_csv}",
        ),
        (
            "SF STATUS = Paid",
            clinical_counts["sf_status_paid"],
            pct_m(clinical_counts["sf_status_paid"]),
            "gap is not missing from billing",
        ),
        (
            "KPI definition unchanged",
            "",
            "",
            "True Match Gap / has_note remains DailyNote-only",
        ),
    ]
    for row in clinical_summary:
        ws6.append(list(row))

    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"Wrote {args.out}", flush=True)
    print(
        {
            "sample_n": len(sample),
            "missing_dn_for_dos": n_miss,
            "has_dn_for_dos": has_dn,
            "lag_1": lag_counts[1],
            "lag_2": lag_counts[2],
            "lag_3": lag_counts[3],
            "lag_1_to_3": lag13_total,
            "wrong_case_exact": wrong_case_exact,
            "wrong_case_lag13": wrong_case_lag13,
            "no_dn_on_case": no_dn_on_case,
            "orphan_note_rows": len(orphan_rows),
            "has_poc_on_case": clinical_counts["has_poc_on_case"],
            "has_discharge_on_case": clinical_counts["has_discharge_on_case"],
            "any_clinical_exact_dos": clinical_counts["any_clinical_exact_dos"],
            "in_sf_billing": clinical_counts["in_sf_billing"],
            "sf_status_paid": clinical_counts["sf_status_paid"],
            "sf_index_size": len(sf_index),
        },
        flush=True,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
