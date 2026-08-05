"""Split unpaid payment rows by whether the patient has an upcoming schedule visit."""

from __future__ import annotations

import argparse
import csv
import sys
from collections import defaultdict
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
REPO_ROOT = ROOT.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import PatternFill  # noqa: E402

from payments_scrape import UNPAID_FIELDNAMES  # noqa: E402

CANCELLED_STATUS = "Cancelled/No Show"
COPAY_TYPE = "copay"
_XLSX_GREEN = "C6EFCE"
_XLSX_YELLOW = "FFEB9C"

CONTACT_FIELDS = (
    "mobile_phone",
    "home_phone",
    "work_phone",
    "email",
    "best_phone",
)

WITH_UPCOMING_FIELDNAMES = [
    *UNPAID_FIELDNAMES,
    "next_appointment_at",
    "next_appointment_status",
    "upcoming_appointment_count",
]

NO_UPCOMING_FIELDNAMES = [
    *UNPAID_FIELDNAMES,
    "has_upcoming",
]

PatientKey = tuple[str, str]


@dataclass(frozen=True)
class UpcomingInfo:
    next_appointment_at: str
    next_appointment_status: str
    upcoming_appointment_count: int


def _patient_key(row: dict[str, Any]) -> PatientKey:
    return (str(row.get("facility_id") or "").strip(), str(row.get("patient_id") or "").strip())


def _appt_day(appointment_at: str) -> date | None:
    day = (appointment_at or "").strip()[:10]
    if len(day) != 10:
        return None
    try:
        return date.fromisoformat(day)
    except ValueError:
        return None


def _month_in_range(iso: str, start_month: str, end_month: str) -> bool:
    ym = (iso or "")[:7]
    return bool(ym) and start_month <= ym <= end_month


def _is_copay(row: dict[str, Any]) -> bool:
    return (row.get("payment_type") or "").strip().lower() == COPAY_TYPE


def filter_copay_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    return [r for r in rows if _is_copay(r)]


def best_phone(mobile: str, home: str, work: str) -> str:
    for value in (mobile, home, work):
        text = (value or "").strip()
        if text:
            return text
    return ""


def load_patient_contacts(patients_csv: Path) -> dict[str, dict[str, str]]:
    """Map PATIENT_ID -> contact fields from patients.csv."""
    out: dict[str, dict[str, str]] = {}
    with patients_csv.open(encoding="utf-8-sig", newline="") as fh:
        for row in csv.DictReader(fh):
            pid = str(row.get("PATIENT_ID") or "").strip()
            if not pid:
                continue
            mobile = str(row.get("MOBILE_PHONE") or "").strip()
            home = str(row.get("HOME_PHONE") or "").strip()
            work = str(row.get("WORK_PHONE") or "").strip()
            email = str(row.get("EMAIL_ADDRESS") or "").strip()
            out[pid] = {
                "mobile_phone": mobile,
                "home_phone": home,
                "work_phone": work,
                "email": email,
                "best_phone": best_phone(mobile, home, work),
            }
    return out


def apply_patient_contacts(
    rows: list[dict[str, Any]],
    contacts: dict[str, dict[str, str]],
) -> None:
    """Fill phone/email fields in place from patients.csv lookup."""
    for row in rows:
        src = contacts.get(str(row.get("patient_id") or "").strip())
        if not src:
            continue
        for field in CONTACT_FIELDS:
            row[field] = src.get(field) or ""


def build_upcoming_by_patient(
    schedule_rows: list[dict[str, str]],
    *,
    as_of: date,
) -> dict[PatientKey, UpcomingInfo]:
    """Patients with ≥1 non-cancelled appointment on/after as_of."""
    by_patient: dict[PatientKey, list[tuple[str, str]]] = defaultdict(list)
    for row in schedule_rows:
        status = (row.get("visit_status") or "").strip()
        if status == CANCELLED_STATUS:
            continue
        at = (row.get("appointment_at") or "").strip()
        day = _appt_day(at)
        if day is None or day < as_of:
            continue
        key = _patient_key(row)
        if not key[1]:
            continue
        by_patient[key].append((at, status))

    out: dict[PatientKey, UpcomingInfo] = {}
    for key, appts in by_patient.items():
        appts.sort(key=lambda t: t[0])
        next_at, next_status = appts[0]
        out[key] = UpcomingInfo(
            next_appointment_at=next_at,
            next_appointment_status=next_status,
            upcoming_appointment_count=len(appts),
        )
    return out


def split_unpaid_by_upcoming(
    unpaid_rows: list[dict[str, str]],
    upcoming: dict[PatientKey, UpcomingInfo],
    *,
    no_upcoming_start_month: str = "2026-01",
    no_upcoming_end_month: str = "2026-05",
    contacts: dict[str, dict[str, str]] | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Return (with_upcoming rows, no_upcoming Jan–May rows). Copay-only."""
    with_upcoming: list[dict[str, Any]] = []
    no_upcoming: list[dict[str, Any]] = []

    for row in filter_copay_rows(unpaid_rows):
        key = _patient_key(row)
        info = upcoming.get(key)
        if info is not None:
            enriched = dict(row)
            enriched["next_appointment_at"] = info.next_appointment_at
            enriched["next_appointment_status"] = info.next_appointment_status
            enriched["upcoming_appointment_count"] = str(info.upcoming_appointment_count)
            with_upcoming.append(enriched)
            continue
        dos = (row.get("dos") or "").strip()
        if not _month_in_range(dos, no_upcoming_start_month, no_upcoming_end_month):
            continue
        out = dict(row)
        out["has_upcoming"] = "0"
        no_upcoming.append(out)

    with_upcoming.sort(
        key=lambda r: (
            str(r.get("facility_id") or ""),
            str(r.get("patient_name") or ""),
            str(r.get("patient_id") or ""),
            str(r.get("dos") or ""),
            str(r.get("next_appointment_at") or ""),
        )
    )
    no_upcoming.sort(
        key=lambda r: (
            str(r.get("facility_id") or ""),
            str(r.get("patient_name") or ""),
            str(r.get("patient_id") or ""),
            str(r.get("dos") or ""),
        )
    )
    if contacts:
        apply_patient_contacts(with_upcoming, contacts)
        apply_patient_contacts(no_upcoming, contacts)
    return with_upcoming, no_upcoming


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as fh:
        return list(csv.DictReader(fh))


def _write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def _write_xlsx(
    path: Path,
    rows: list[dict[str, Any]],
    fieldnames: list[str],
    *,
    sheet_title: str,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]
    ws.append(fieldnames)

    counts: dict[PatientKey, int] = defaultdict(int)
    for r in rows:
        counts[_patient_key(r)] += 1
    seen: dict[PatientKey, int] = defaultdict(int)
    green = PatternFill(start_color=_XLSX_GREEN, end_color=_XLSX_GREEN, fill_type="solid")
    yellow = PatternFill(
        start_color=_XLSX_YELLOW, end_color=_XLSX_YELLOW, fill_type="solid"
    )

    for r in rows:
        ws.append([r.get(c, "") for c in fieldnames])
        key = _patient_key(r)
        if counts[key] <= 1:
            continue
        seen[key] += 1
        fill = green if seen[key] == 1 else yellow
        for col in range(1, len(fieldnames) + 1):
            ws.cell(row=ws.max_row, column=col).fill = fill

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _write_combined_xlsx(
    path: Path,
    *,
    with_upcoming: list[dict[str, Any]],
    no_upcoming: list[dict[str, Any]],
) -> None:
    wb = Workbook()
    sheets = (
        ("with upcoming", with_upcoming, WITH_UPCOMING_FIELDNAMES),
        ("no upcoming Jan-May", no_upcoming, NO_UPCOMING_FIELDNAMES),
    )
    for idx, (title, rows, fields) in enumerate(sheets):
        ws = wb.active if idx == 0 else wb.create_sheet()
        ws.title = title[:31]
        ws.append(fields)
        counts: dict[PatientKey, int] = defaultdict(int)
        for r in rows:
            counts[_patient_key(r)] += 1
        seen: dict[PatientKey, int] = defaultdict(int)
        green = PatternFill(
            start_color=_XLSX_GREEN, end_color=_XLSX_GREEN, fill_type="solid"
        )
        yellow = PatternFill(
            start_color=_XLSX_YELLOW, end_color=_XLSX_YELLOW, fill_type="solid"
        )
        for r in rows:
            ws.append([r.get(c, "") for c in fields])
            key = _patient_key(r)
            if counts[key] <= 1:
                continue
            seen[key] += 1
            fill = green if seen[key] == 1 else yellow
            for col in range(1, len(fields) + 1):
                ws.cell(row=ws.max_row, column=col).fill = fill

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def run_split(
    *,
    unpaid_csv: Path,
    schedule_csv: Path,
    patients_csv: Path,
    output_dir: Path,
    as_of: date,
    no_upcoming_start_month: str,
    no_upcoming_end_month: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    unpaid_rows = _read_csv(unpaid_csv)
    schedule_rows = _read_csv(schedule_csv)
    contacts = load_patient_contacts(patients_csv)
    upcoming = build_upcoming_by_patient(schedule_rows, as_of=as_of)
    with_upcoming, no_upcoming = split_unpaid_by_upcoming(
        unpaid_rows,
        upcoming,
        no_upcoming_start_month=no_upcoming_start_month,
        no_upcoming_end_month=no_upcoming_end_month,
        contacts=contacts,
    )

    tag = as_of.isoformat()
    with_csv = output_dir / f"missing_payments_with_upcoming_asof_{tag}.csv"
    with_xlsx = output_dir / f"missing_payments_with_upcoming_asof_{tag}.xlsx"
    no_csv = output_dir / f"missing_payments_no_upcoming_jan_may_asof_{tag}.csv"
    no_xlsx = output_dir / f"missing_payments_no_upcoming_jan_may_asof_{tag}.xlsx"
    combined = output_dir / f"missing_payments_by_upcoming_asof_{tag}.xlsx"

    _write_csv(with_csv, with_upcoming, WITH_UPCOMING_FIELDNAMES)
    _write_xlsx(with_xlsx, with_upcoming, WITH_UPCOMING_FIELDNAMES, sheet_title="with upcoming")
    _write_csv(no_csv, no_upcoming, NO_UPCOMING_FIELDNAMES)
    _write_xlsx(
        no_xlsx, no_upcoming, NO_UPCOMING_FIELDNAMES, sheet_title="no upcoming Jan-May"
    )
    _write_combined_xlsx(
        combined, with_upcoming=with_upcoming, no_upcoming=no_upcoming
    )

    with_contact = sum(1 for r in with_upcoming if r.get("best_phone") or r.get("email"))
    no_contact = sum(1 for r in no_upcoming if r.get("best_phone") or r.get("email"))
    print(f"as_of={tag} payment_type=Copay patients_csv={patients_csv}")
    print(f"upcoming patients in schedule: {len(upcoming)}")
    print(
        f"with upcoming: {len(with_upcoming)} rows / "
        f"{len({_patient_key(r) for r in with_upcoming})} patients "
        f"(with phone/email: {with_contact}) -> {with_csv}"
    )
    print(
        f"no upcoming ({no_upcoming_start_month}..{no_upcoming_end_month}): "
        f"{len(no_upcoming)} rows / "
        f"{len({_patient_key(r) for r in no_upcoming})} patients "
        f"(with phone/email: {no_contact}) -> {no_csv}"
    )
    print(f"combined xlsx -> {combined}")
    return with_upcoming, no_upcoming


def main() -> None:
    default_out = ROOT / "output" / "jan_aug_2026"
    p = argparse.ArgumentParser(
        description="Split unpaid Copay payments by upcoming schedule appointments"
    )
    p.add_argument(
        "--unpaid-csv",
        type=Path,
        default=default_out / "patient_payments_unpaid_202601_202608.csv",
    )
    p.add_argument(
        "--schedule-csv",
        type=Path,
        default=default_out / "schedule_visits_2026-01-01_2026-08-30.csv",
    )
    p.add_argument(
        "--patients-csv",
        type=Path,
        default=REPO_ROOT / "patients.csv",
    )
    p.add_argument("--output-dir", type=Path, default=default_out)
    p.add_argument("--as-of", type=str, default="2026-07-28", help="YYYY-MM-DD inclusive")
    p.add_argument("--no-upcoming-start-month", type=str, default="2026-01")
    p.add_argument("--no-upcoming-end-month", type=str, default="2026-05")
    args = p.parse_args()

    run_split(
        unpaid_csv=args.unpaid_csv,
        schedule_csv=args.schedule_csv,
        patients_csv=args.patients_csv,
        output_dir=args.output_dir,
        as_of=date.fromisoformat(args.as_of),
        no_upcoming_start_month=args.no_upcoming_start_month,
        no_upcoming_end_month=args.no_upcoming_end_month,
    )


if __name__ == "__main__":
    main()
