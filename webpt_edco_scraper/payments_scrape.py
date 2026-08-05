"""CLI logic: scrape Patient Payments for a cohort and build unpaid sheet."""

from __future__ import annotations

import asyncio
import csv
import json
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from playwright.async_api import async_playwright

from auth import (
    ClinicSwitchError,
    create_context,
    ensure_authenticated,
    list_clinics,
    safe_close_context,
    save_storage_state,
    switch_clinic_and_settle,
)
from config import STORAGE_STATE_PATH, WebPTConfig
from logging_config import get_logger
from patient_payments_api import (
    fetch_patient_payments,
    parse_mmddyyyy,
)

log = get_logger("payments_scrape")

PAYMENT_FIELDNAMES = [
    "facility_id",
    "facility_name",
    "patient_id",
    "patient_name",
    "case_id",
    "date_of_service",
    "date_of_service_iso",
    "date_of_transaction",
    "payment_type",
    "description",
    "amount_due",
    "amount_paid",
    "paid_method",
    "credit_type",
    "auth_check",
    "total_charge",
    "total_paid",
    "balance",
]

UNPAID_FIELDNAMES = [
    "facility_id",
    "facility_name",
    "patient_id",
    "patient_name",
    "case_id",
    "dos",
    "payment_type",
    "description",
    "amount_due",
    "amount_paid",
    "amount_owed",
    "reason",
    "mobile_phone",
    "home_phone",
    "work_phone",
    "email",
    "best_phone",
]

_XLSX_GREEN = "C6EFCE"
_XLSX_YELLOW = "FFEB9C"


def _month_in_range(iso: str, start: str, end: str) -> bool:
    # iso YYYY-MM-DD; start/end YYYY-MM
    ym = iso[:7]
    return start <= ym <= end


def _parse_appt_dates(row: dict[str, str]) -> list[str]:
    out: list[str] = []
    for col in (
        "appointment_dates",
        "appointments_past_dates",
        "appointments_upcoming_dates",
    ):
        for part in (row.get(col) or "").split(";"):
            part = part.strip()
            if not part:
                continue
            out.append(part[:10])
    return out


def build_outreach_jan_may_cohort(
    *,
    outreach_csv: Path,
    export_csv: Path,
    start_month: str = "2026-01",
    end_month: str = "2026-05",
) -> list[dict[str, str]]:
    export_by_key: dict[tuple[str, str], dict[str, str]] = {}
    with export_csv.open(encoding="utf-8-sig", newline="") as fh:
        for r in csv.DictReader(fh):
            export_by_key[(str(r.get("facility_id") or ""), str(r.get("patient_id") or ""))] = r

    cohort: list[dict[str, str]] = []
    with outreach_csv.open(encoding="utf-8-sig", newline="") as fh:
        for r in csv.DictReader(fh):
            key = (str(r.get("facility_id") or ""), str(r.get("patient_id") or ""))
            src = export_by_key.get(key)
            if not src:
                continue
            dates = _parse_appt_dates(src)
            if not any(
                len(d) >= 7 and start_month <= d[:7] <= end_month for d in dates
            ):
                continue
            merged = dict(r)
            merged["appointment_dates_jan_may"] = "; ".join(
                d for d in dates if len(d) >= 7 and start_month <= d[:7] <= end_month
            )
            # Prefer case_id from export if outreach missing
            if not (merged.get("case_id") or "").strip():
                merged["case_id"] = src.get("case_id") or ""
            cohort.append(merged)

    cohort.sort(
        key=lambda r: (
            str(r.get("facility_id") or ""),
            str(r.get("patient_id") or ""),
        )
    )
    return cohort


def build_export_payments_cohort(
    *,
    export_csv: Path,
    start_month: str = "2026-01",
    end_month: str = "2026-08",
) -> list[dict[str, str]]:
    """All patients in export with a case_id (optionally filtered by appt month)."""
    cohort: list[dict[str, str]] = []
    seen: set[tuple[str, str, str]] = set()
    with export_csv.open(encoding="utf-8-sig", newline="") as fh:
        for r in csv.DictReader(fh):
            fid = str(r.get("facility_id") or "").strip()
            pid = str(r.get("patient_id") or "").strip()
            cid = str(r.get("case_id") or "").strip()
            if not pid or not cid:
                continue
            dates = _parse_appt_dates(r)
            if dates and not any(
                len(d) >= 7 and start_month <= d[:7] <= end_month for d in dates
            ):
                continue
            key = (fid, pid, cid)
            if key in seen:
                continue
            seen.add(key)
            cohort.append(
                {
                    "facility_id": fid,
                    "facility_name": str(r.get("facility_name") or ""),
                    "patient_id": pid,
                    "patient_name": str(r.get("patient_name") or ""),
                    "case_id": cid,
                    "mobile_phone": str(r.get("mobile_phone") or ""),
                    "home_phone": str(r.get("home_phone") or ""),
                    "work_phone": str(r.get("work_phone") or ""),
                    "email": str(r.get("email") or ""),
                    "best_phone": str(r.get("best_phone") or ""),
                }
            )
    cohort.sort(
        key=lambda row: (
            str(row.get("facility_id") or ""),
            str(row.get("patient_id") or ""),
            str(row.get("case_id") or ""),
        )
    )
    return cohort


def _load_checkpoint(path: Path) -> set[str]:
    if not path.exists():
        return set()
    data = json.loads(path.read_text(encoding="utf-8"))
    return set(data.get("done_keys") or [])


def _save_checkpoint(path: Path, done: set[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps({"done_keys": sorted(done)}, indent=2),
        encoding="utf-8",
    )


def _write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def _append_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    """Append rows; write header only when the file is new/empty."""
    if not rows:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    write_header = not path.exists() or path.stat().st_size == 0
    with path.open("a", encoding="utf-8", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=fieldnames, extrasaction="ignore")
        if write_header:
            w.writeheader()
        w.writerows(rows)


def _write_xlsx(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    from openpyxl.styles import PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "unpaid"
    ws.append(fieldnames)

    # Highlight repeat patients: first row green, later rows yellow.
    counts: dict[tuple[str, str], int] = defaultdict(int)
    for r in rows:
        counts[(str(r.get("facility_id") or ""), str(r.get("patient_id") or ""))] += 1
    seen: dict[tuple[str, str], int] = defaultdict(int)
    green = PatternFill(start_color=_XLSX_GREEN, end_color=_XLSX_GREEN, fill_type="solid")
    yellow = PatternFill(
        start_color=_XLSX_YELLOW, end_color=_XLSX_YELLOW, fill_type="solid"
    )

    for r in rows:
        ws.append([r.get(c, "") for c in fieldnames])
        key = (str(r.get("facility_id") or ""), str(r.get("patient_id") or ""))
        if counts[key] <= 1:
            continue
        seen[key] += 1
        fill = green if seen[key] == 1 else yellow
        for col in range(1, len(fieldnames) + 1):
            ws.cell(row=ws.max_row, column=col).fill = fill

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def build_unpaid_rows(
    *,
    payment_rows: list[dict[str, Any]],
    cohort: list[dict[str, str]],
    start_month: str,
    end_month: str,
) -> list[dict[str, Any]]:
    """Build underpaid-only outreach rows (amount_paid < amount_due)."""
    unpaid: list[dict[str, Any]] = []
    cohort_by_pid = {str(r.get("patient_id") or ""): r for r in cohort}

    for p in payment_rows:
        iso = p.get("date_of_service_iso") or ""
        if not iso or not _month_in_range(iso, start_month, end_month):
            continue
        due = float(p.get("amount_due") or 0)
        paid = float(p.get("amount_paid") or 0)
        if due <= paid + 0.009:
            continue
        pid = str(p.get("patient_id") or "")
        owed = due - paid
        unpaid.append(
            {
                "facility_id": p.get("facility_id") or "",
                "facility_name": p.get("facility_name") or "",
                "patient_id": pid,
                "patient_name": p.get("patient_name") or "",
                "case_id": p.get("case_id") or "",
                "dos": iso,
                "payment_type": p.get("payment_type") or "",
                "description": p.get("description") or "",
                "amount_due": f"{due:.2f}",
                "amount_paid": f"{paid:.2f}",
                "amount_owed": f"{owed:.2f}",
                "reason": "underpaid",
                "mobile_phone": "",
                "home_phone": "",
                "work_phone": "",
                "email": "",
                "best_phone": "",
            }
        )

    for u in unpaid:
        src = cohort_by_pid.get(str(u.get("patient_id") or ""))
        if not src:
            continue
        for k in (
            "mobile_phone",
            "home_phone",
            "work_phone",
            "email",
            "best_phone",
            "facility_name",
        ):
            if not u.get(k):
                u[k] = src.get(k) or ""
        if not u.get("facility_id"):
            u["facility_id"] = src.get("facility_id") or ""

    unpaid.sort(
        key=lambda r: (
            str(r.get("facility_id") or ""),
            str(r.get("patient_name") or ""),
            str(r.get("patient_id") or ""),
            str(r.get("dos") or ""),
        )
    )
    return unpaid


def rebuild_unpaid_exports(
    *,
    payments_csv: Path,
    outreach_csv: Path,
    export_csv: Path,
    unpaid_csv: Path,
    unpaid_xlsx: Path,
    start_month: str = "2026-01",
    end_month: str = "2026-05",
) -> list[dict[str, Any]]:
    """Rebuild underpaid CSV/XLSX from an existing payments dump (no scrape)."""
    with payments_csv.open(encoding="utf-8-sig", newline="") as fh:
        payment_rows = list(csv.DictReader(fh))
    cohort = build_outreach_jan_may_cohort(
        outreach_csv=outreach_csv,
        export_csv=export_csv,
        start_month=start_month,
        end_month=end_month,
    )
    unpaid = build_unpaid_rows(
        payment_rows=payment_rows,
        cohort=cohort,
        start_month=start_month,
        end_month=end_month,
    )
    _write_csv(unpaid_csv, unpaid, UNPAID_FIELDNAMES)
    _write_xlsx(unpaid_xlsx, unpaid, UNPAID_FIELDNAMES)
    log.info(
        "Rebuilt underpaid-only unpaid=%s (%d) xlsx=%s",
        unpaid_csv,
        len(unpaid),
        unpaid_xlsx,
    )
    return unpaid


async def cmd_scrape_patient_payments(
    config: WebPTConfig,
    *,
    outreach_csv: Path | None,
    export_csv: Path,
    output_dir: Path,
    start_month: str = "2026-01",
    end_month: str = "2026-05",
    max_patients: int | None = None,
    concurrency: int = 10,
    all_export: bool = False,
    assert_exclusive=None,
) -> None:
    if assert_exclusive is not None:
        assert_exclusive()

    if all_export or outreach_csv is None:
        cohort = build_export_payments_cohort(
            export_csv=export_csv,
            start_month=start_month,
            end_month=end_month,
        )
        log.info(
            "Payments cohort export-all %s..%s: %d patients",
            start_month,
            end_month,
            len(cohort),
        )
    else:
        cohort = build_outreach_jan_may_cohort(
            outreach_csv=outreach_csv,
            export_csv=export_csv,
            start_month=start_month,
            end_month=end_month,
        )
        log.info(
            "Payments cohort outreach∩%s..%s: %d patients",
            start_month,
            end_month,
            len(cohort),
        )
    if max_patients is not None:
        cohort = cohort[: max(0, max_patients)]
        log.info("Capped to max_patients=%d", len(cohort))

    output_dir.mkdir(parents=True, exist_ok=True)
    debug_dir = output_dir / "debug"
    checkpoint_path = output_dir / "payments_checkpoint.json"
    tag = f"{start_month.replace('-', '')}_{end_month.replace('-', '')}"
    payments_path = output_dir / f"patient_payments_{tag}.csv"
    unpaid_csv = output_dir / f"patient_payments_unpaid_{tag}.csv"
    unpaid_xlsx = output_dir / f"patient_payments_unpaid_{tag}.xlsx"

    done = _load_checkpoint(checkpoint_path)
    # Load existing payments if resuming
    payment_rows: list[dict[str, Any]] = []
    if payments_path.exists():
        with payments_path.open(encoding="utf-8-sig", newline="") as fh:
            payment_rows = list(csv.DictReader(fh))

    pending = [
        r
        for r in cohort
        if f"{r.get('facility_id')}:{r.get('patient_id')}:{r.get('case_id')}"
        not in done
    ]
    log.info("Pending %d / %d (checkpoint done=%d)", len(pending), len(cohort), len(done))

    async with async_playwright() as playwright:
        context = await create_context(playwright, config)
        page = await context.new_page()
        try:
            await ensure_authenticated(page, context, config)
            clinics = await list_clinics(page, config.company_id)
            clinics_by_id = {str(c.facility_id): c for c in clinics}
            sem = asyncio.Semaphore(max(1, concurrency))
            current_facility: str | None = None

            # Group pending by facility for settle-once
            by_fac: dict[str, list[dict[str, str]]] = defaultdict(list)
            for r in pending:
                by_fac[str(r.get("facility_id") or "")].append(r)

            for fid in sorted(by_fac.keys()):
                rows = by_fac[fid]
                clinic = clinics_by_id.get(fid)
                if clinic is None:
                    log.warning("Unknown facility %s — skipping %d patients", fid, len(rows))
                    continue
                if fid != current_facility:
                    settled = False
                    for attempt in range(3):
                        try:
                            await switch_clinic_and_settle(
                                page,
                                context,
                                config,
                                company_id=clinic.company_id,
                                facility_id=clinic.facility_id,
                            )
                            settled = True
                            break
                        except ClinicSwitchError as exc:
                            log.warning(
                                "Clinic switch failed facility=%s attempt=%d: %s",
                                fid,
                                attempt + 1,
                                exc,
                            )
                            await ensure_authenticated(page, context, config)
                            await asyncio.sleep(2)
                    if not settled:
                        log.error(
                            "Skipping facility %s (%d patients) — clinic switch failed",
                            fid,
                            len(rows),
                        )
                        continue
                    current_facility = fid
                    log.info("Settled facility %s (%d patients)", fid, len(rows))
                    # Warm HTTP session with one sequential page fetch so concurrent
                    # context.request calls see an authenticated clinic cookie.
                    warm = rows[0]
                    try:
                        warm_pid = int(warm.get("patient_id") or 0)
                        warm_cid = int(warm.get("case_id") or 0)
                        if warm_pid and warm_cid:
                            await fetch_patient_payments(
                                context,
                                patient_id=warm_pid,
                                case_id=warm_cid,
                                page=page,
                                config=config,
                                prefer_http=False,
                                debug_dir=debug_dir,
                                timeout_ms=30_000,
                                retries=1,
                            )
                    except Exception as warm_exc:  # noqa: BLE001
                        log.warning(
                            "Payments warm-up failed facility=%s: %s",
                            fid,
                            warm_exc,
                        )

                async def _one(row: dict[str, str]) -> tuple[str, list[dict[str, Any]], str]:
                    pid = int(row.get("patient_id") or 0)
                    cid_raw = row.get("case_id") or ""
                    if not cid_raw.strip():
                        key = f"{row.get('facility_id')}:{pid}:"
                        return key, [], "missing_case_id"
                    cid = int(cid_raw)
                    key = f"{row.get('facility_id')}:{pid}:{cid}"
                    # page=None: concurrent workers must not share page.goto fallbacks
                    async with sem:
                        try:
                            result = await asyncio.wait_for(
                                fetch_patient_payments(
                                    context,
                                    patient_id=pid,
                                    case_id=cid,
                                    page=None,
                                    config=config,
                                    prefer_http=True,
                                    debug_dir=debug_dir,
                                    timeout_ms=25_000,
                                    retries=2,
                                ),
                                timeout=40.0,
                            )
                        except asyncio.TimeoutError:
                            return key, [], "timeout"
                    out_rows: list[dict[str, Any]] = []
                    for txn in result.transactions:
                        iso = parse_mmddyyyy(txn.date_of_service)
                        if not iso or not _month_in_range(iso, start_month, end_month):
                            continue
                        out_rows.append(
                            {
                                "facility_id": row.get("facility_id") or "",
                                "facility_name": row.get("facility_name") or "",
                                "patient_id": str(pid),
                                "patient_name": row.get("patient_name") or "",
                                "case_id": str(cid),
                                "date_of_service": txn.date_of_service,
                                "date_of_service_iso": iso,
                                "date_of_transaction": txn.date_of_transaction,
                                "payment_type": txn.payment_type,
                                "description": txn.description,
                                "amount_due": f"{txn.amount_due:.2f}",
                                "amount_paid": f"{txn.amount_paid:.2f}",
                                "paid_method": txn.paid_method,
                                "credit_type": txn.credit_type,
                                "auth_check": txn.auth_check,
                                "total_charge": f"{result.total_charge:.2f}",
                                "total_paid": f"{result.total_paid:.2f}",
                                "balance": f"{result.balance:.2f}",
                            }
                        )
                    return key, out_rows, result.fetch_error

                # Process in batches to checkpoint
                batch_size = max(concurrency * 2, 20)
                for i in range(0, len(rows), batch_size):
                    batch = rows[i : i + batch_size]
                    results = await asyncio.gather(*[_one(r) for r in batch])
                    batch_new: list[dict[str, Any]] = []
                    for key, out_rows, err in results:
                        if err and err != "missing_case_id":
                            log.warning(
                                "Payments %s error=%s rows=%d (not checkpointed)",
                                key,
                                err,
                                len(out_rows),
                            )
                            continue
                        done.add(key)
                        payment_rows.extend(out_rows)
                        batch_new.extend(out_rows)
                        if err:
                            log.warning(
                                "Payments %s error=%s rows=%d",
                                key,
                                err,
                                len(out_rows),
                            )
                    _save_checkpoint(checkpoint_path, done)
                    _append_csv(payments_path, batch_new, PAYMENT_FIELDNAMES)
                    log.info(
                        "Facility %s progress %d/%d | payments_rows=%d done=%d",
                        fid,
                        min(i + batch_size, len(rows)),
                        len(rows),
                        len(payment_rows),
                        len(done),
                    )

            await save_storage_state(context, STORAGE_STATE_PATH)
        finally:
            await safe_close_context(context)

    unpaid = build_unpaid_rows(
        payment_rows=payment_rows,
        cohort=cohort,
        start_month=start_month,
        end_month=end_month,
    )
    _write_csv(unpaid_csv, unpaid, UNPAID_FIELDNAMES)
    _write_xlsx(unpaid_xlsx, unpaid, UNPAID_FIELDNAMES)
    log.info(
        "Wrote payments=%s (%d) unpaid=%s (%d)",
        payments_path,
        len(payment_rows),
        unpaid_csv,
        len(unpaid),
    )
