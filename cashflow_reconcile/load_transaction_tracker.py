"""Load bank deposit dates from Transaction Tracker (Postgres SoT, xlsx fallback)."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from .normalize import parse_date
from .payer_registry import (
    extract_ach_payer_head,
    extract_eft_refs_from_description,
    resolve_tracker_description,
)

MONTH_SHEETS = frozenset(
    {
        "January",
        "February",
        "March",
        "April",
        "May",
        "June",
        "July",
        "August",
        "September",
        "October",
        "November",
        "December",
    }
)

EFT_COLUMNS = ("EFT_1", "EFT_2", "#Check/Reference")
SKIP_VALUES = frozenset({"", "#N/A", "N/A", "NONE", "NULL"})


def normalize_eft(value: object) -> str:
    """Normalize check/EFT reference text for tracker lookups."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        text = str(int(value))
    else:
        text = str(value).strip()
    if not text or text.upper() in SKIP_VALUES:
        return ""
    return text


# Backwards-compatible alias used inside this module.
_normalize_eft = normalize_eft


def _coerce_date(value: object) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return parse_date(str(value))


def _format_deposit_date(value: date) -> str:
    return value.strftime("%m/%d/%Y")


def _coerce_amount(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    from decimal import Decimal

    if isinstance(value, Decimal):
        return float(value)
    text = str(value).strip().replace(",", "").replace("$", "")
    if not text or text.upper() in SKIP_VALUES:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _payer_fields_from_description(description: str) -> dict[str, str]:
    head = extract_ach_payer_head(description)
    resolved = resolve_tracker_description(description)
    return {
        "ach_payer_head": head,
        "payer_org_code": resolved.code if resolved else "",
        "payer_org": resolved.name if resolved else "",
    }


def _sheet_label_from_date(d: date | None) -> str:
    if d is None:
        return ""
    import calendar

    return calendar.month_name[d.month]


def _ledger_from_db_rows(db_rows: list[dict]) -> list[dict[str, object]]:
    rows_out: list[dict[str, object]] = []
    for raw in db_rows:
        deposit = _coerce_date(raw.get("txn_date"))
        amount = _coerce_amount(raw.get("amount"))
        if deposit is None or amount is None:
            continue
        description = str(raw.get("description") or "").strip()
        payer_fields = _payer_fields_from_description(description)
        eft_1 = _normalize_eft(raw.get("eft_1"))
        eft_2 = _normalize_eft(raw.get("eft_2"))
        check_reference = _normalize_eft(raw.get("check_reference"))
        for ref in extract_eft_refs_from_description(description):
            norm = _normalize_eft(ref)
            if not norm:
                continue
            if not eft_1:
                eft_1 = norm
            elif norm not in {eft_1, eft_2, check_reference} and not eft_2:
                eft_2 = norm
            elif norm not in {eft_1, eft_2, check_reference} and not check_reference:
                check_reference = norm
        rows_out.append(
            {
                "payment_id": str(raw.get("payment_id") or "").strip(),
                "deposit_date": deposit,
                "amount": amount,
                "eft_1": eft_1,
                "eft_2": eft_2,
                "check_reference": check_reference,
                "description": description,
                "sheet": _sheet_label_from_date(deposit),
                **payer_fields,
            }
        )
    return rows_out


def _dates_from_db_rows(db_rows: list[dict]) -> dict[str, str]:
    earliest: dict[str, date] = {}
    for raw in db_rows:
        deposit = _coerce_date(raw.get("txn_date"))
        if deposit is None:
            continue
        description = str(raw.get("description") or "").strip()
        keys: list[str] = []
        for value in (raw.get("eft_1"), raw.get("eft_2"), raw.get("check_reference")):
            key = _normalize_eft(value)
            if key:
                keys.append(key)
        for ref in extract_eft_refs_from_description(description):
            key = _normalize_eft(ref)
            if key and key not in keys:
                keys.append(key)
        for key in keys:
            current = earliest.get(key)
            if current is None or deposit < current:
                earliest[key] = deposit
    return {key: _format_deposit_date(value) for key, value in earliest.items()}


def load_deposit_ledger(path: Path | None = None) -> list[dict[str, object]]:
    """Load bank deposit rows: deposit_date, amount, eft refs, payment_id.

    Postgres active tracker rows are the default source. Pass an xlsx ``path``
    only for fixtures / one-off file reads.
    """
    if path is not None:
        return _load_deposit_ledger_xlsx(path)

    from cashflow_db.repository import connection, tracker

    with connection() as conn:
        db_rows = tracker.list_active_for_etl(conn)
    return _ledger_from_db_rows(db_rows)


def _load_deposit_ledger_xlsx(path: Path) -> list[dict[str, object]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    rows_out: list[dict[str, object]] = []
    try:
        for sheet_name in workbook.sheetnames:
            if sheet_name not in MONTH_SHEETS:
                continue
            worksheet = workbook[sheet_name]
            header: list[str] | None = None
            for index, row in enumerate(worksheet.iter_rows(values_only=True)):
                if index == 0:
                    header = [
                        str(cell).strip() if cell is not None else "" for cell in row
                    ]
                    continue
                if header is None or not row or row[0] is None:
                    continue
                record = dict(zip(header, row))
                deposit = _coerce_date(record.get("Date"))
                amount = _coerce_amount(record.get("Amount"))
                if deposit is None or amount is None:
                    continue
                description = str(record.get("Description") or "").strip()
                payer_fields = _payer_fields_from_description(description)
                eft_1 = _normalize_eft(record.get("EFT_1"))
                eft_2 = _normalize_eft(record.get("EFT_2"))
                check_reference = _normalize_eft(record.get("#Check/Reference"))
                for ref in extract_eft_refs_from_description(description):
                    norm = _normalize_eft(ref)
                    if not norm:
                        continue
                    if not eft_1:
                        eft_1 = norm
                    elif norm not in {eft_1, eft_2, check_reference} and not eft_2:
                        eft_2 = norm
                    elif (
                        norm not in {eft_1, eft_2, check_reference}
                        and not check_reference
                    ):
                        check_reference = norm
                rows_out.append(
                    {
                        "payment_id": str(record.get("Payment ID") or "").strip(),
                        "deposit_date": deposit,
                        "amount": amount,
                        "eft_1": eft_1,
                        "eft_2": eft_2,
                        "check_reference": check_reference,
                        "description": description,
                        "sheet": sheet_name,
                        **payer_fields,
                    }
                )
    finally:
        workbook.close()
    return rows_out


def load_deposit_dates(path: Path | None = None) -> dict[str, str]:
    """Map check/EFT numbers to bank deposit dates (MM/DD/YYYY).

    Postgres active tracker rows are the default source. Pass an xlsx ``path``
    only for fixtures / one-off file reads.
    """
    if path is not None:
        return _load_deposit_dates_xlsx(path)

    from cashflow_db.repository import connection, tracker

    with connection() as conn:
        db_rows = tracker.list_active_for_etl(conn)
    return _dates_from_db_rows(db_rows)


def _load_deposit_dates_xlsx(path: Path) -> dict[str, str]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    earliest: dict[str, date] = {}
    try:
        for sheet_name in workbook.sheetnames:
            if sheet_name not in MONTH_SHEETS:
                continue
            worksheet = workbook[sheet_name]
            header: list[str] | None = None
            for index, row in enumerate(worksheet.iter_rows(values_only=True)):
                if index == 0:
                    header = [
                        str(cell).strip() if cell is not None else "" for cell in row
                    ]
                    continue
                if header is None or not row or row[0] is None:
                    continue
                record = dict(zip(header, row))
                deposit = _coerce_date(record.get("Date"))
                if deposit is None:
                    continue
                description = str(record.get("Description") or "").strip()
                keys: list[str] = []
                for column in EFT_COLUMNS:
                    key = _normalize_eft(record.get(column))
                    if key:
                        keys.append(key)
                for ref in extract_eft_refs_from_description(description):
                    key = _normalize_eft(ref)
                    if key and key not in keys:
                        keys.append(key)
                for key in keys:
                    current = earliest.get(key)
                    if current is None or deposit < current:
                        earliest[key] = deposit
    finally:
        workbook.close()

    return {key: _format_deposit_date(value) for key, value in earliest.items()}


def apply_deposit_dates(
    checks: dict[str, dict],
    deposit_dates: dict[str, str] | None,
) -> None:
    """Override rollup check dates in-place when a bank deposit date is known."""
    if not deposit_dates:
        return
    for check_num, meta in checks.items():
        bank_date = deposit_dates.get(check_num)
        if bank_date:
            meta["date"] = bank_date
