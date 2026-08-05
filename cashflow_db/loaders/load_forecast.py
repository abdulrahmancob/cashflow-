"""Load rules seed extras + versioned forecast predictions."""

from __future__ import annotations

import csv
import json
from datetime import date, timedelta
from pathlib import Path

from cashflow_db.config import (
    ICD_DENIAL_XLSX,
    PAYABLE_CPT_CSV,
    WEBPT_LEGACY_OUTPUT,
    WEBPT_OUTPUT,
)
from cashflow_db.db import connect, finish_etl_run, start_etl_run
from cashflow_db.loaders.base import load_business_rules
from cashflow_db.util import parse_date, parse_money, safe_str

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None


def load_rules(*, database_url: str | None = None) -> dict[str, int]:
    """Load ICD denial matrix + payable CPT preferences into ref tables."""
    counts = {"icd_rules": 0, "cpt_rules": 0}
    with connect(database_url) as conn:
        etl_id = start_etl_run(conn, "rules", str(ICD_DENIAL_XLSX))
        try:
            if openpyxl and ICD_DENIAL_XLSX.exists():
                wb = openpyxl.load_workbook(ICD_DENIAL_XLSX, read_only=True, data_only=True)
                if "Denial Matrix" in wb.sheetnames:
                    ws = wb["Denial Matrix"]
                    rows = ws.iter_rows(values_only=True)
                    next(rows, None)
                    for row in rows:
                        if not row or not row[0]:
                            continue
                        conn.execute(
                            """
                            INSERT INTO ref.icd_denial_rule (
                                category, description, examples, correct_approach, severity
                            )
                            VALUES (%s, %s, %s, %s, 'error')
                            """,
                            (
                                safe_str(row[0]),
                                safe_str(row[1]) if len(row) > 1 else None,
                                safe_str(row[2]) if len(row) > 2 else None,
                                safe_str(row[3]) if len(row) > 3 else None,
                            ),
                        )
                        counts["icd_rules"] += 1
                wb.close()

            if PAYABLE_CPT_CSV.exists():
                with PAYABLE_CPT_CSV.open("r", encoding="utf-8-sig", errors="replace", newline="") as fh:
                    reader = csv.DictReader(fh)
                    for row in reader:
                        ins = safe_str(row.get("Insurance"))
                        if not ins:
                            continue
                        # store as unbound payer_cpt_rule (plan nullable)
                        for kind, col in (
                            ("preferred", "Preferred"),
                            ("highly_preferred", "Highly preferred"),
                            ("do_not_use", "Do not use"),
                            ("e_stim", "Accepted E-stim code"),
                            ("use", "Use"),
                        ):
                            val = safe_str(row.get(col))
                            if not val or val == "-":
                                continue
                            conn.execute(
                                """
                                INSERT INTO ref.payer_cpt_rule (
                                    insurance_plan_id, rule_kind, expected_value, severity, detail
                                )
                                VALUES (NULL, %s, %s, 'warning', %s)
                                """,
                                (kind, val, ins),
                            )
                            counts["cpt_rules"] += 1

            finish_etl_run(conn, etl_id, status="success", row_count=sum(counts.values()))
        except Exception as exc:
            finish_etl_run(conn, etl_id, status="failed", notes=str(exc)[:2000])
            raise
    return counts


def load_forecast_from_csv(
    *,
    root: Path | None = None,
    database_url: str | None = None,
    algorithm_version: str | None = None,
) -> dict[str, int]:
    """Import existing outcome_stages.csv as an immutable forecast_run."""
    if root is None:
        for candidate in (WEBPT_OUTPUT, WEBPT_LEGACY_OUTPUT):
            if (candidate / "forecast" / "outcome_stages.csv").exists():
                root = candidate
                break
        else:
            root = WEBPT_OUTPUT
    path = root / "forecast" / "outcome_stages.csv"
    rules = load_business_rules()
    defaults = rules.get("forecast_defaults", {})
    algorithm_version = algorithm_version or defaults.get("algorithm_version", "v1-csv-import")
    params = {
        "first_pass_target": defaults.get("first_pass_target", 0.8),
        "denial_shift_cycles": defaults.get("denial_shift_cycles", 1),
        "medical_audit_delay_pct": defaults.get("medical_audit_delay_pct", 0.05),
        "auth_delay_pct": defaults.get("auth_delay_pct", 0.05),
        "high_season_months": defaults.get("high_season_months", [10, 11, 12]),
        "source_file": str(path),
    }
    counts = {"predictions": 0}

    with connect(database_url) as conn:
        etl_id = start_etl_run(conn, "forecast", str(path))
        try:
            run = conn.execute(
                """
                INSERT INTO analytics.forecast_run (
                    algorithm_version, params, as_of_date, status, etl_run_id
                )
                VALUES (%s, %s::jsonb, %s, 'success', %s::uuid)
                RETURNING forecast_run_id
                """,
                (algorithm_version, json.dumps(params), date.today(), etl_id),
            ).fetchone()
            run_id = str(run["forecast_run_id"])

            if not path.exists():
                # synthesize from claim_lines if CSV missing
                rows = conn.execute(
                    """
                    SELECT cl.claim_line_id, cl.visit_id, cl.expected_amount, cl.billed_amount,
                           v.service_date, c.status_current
                    FROM billing.claim_line cl
                    JOIN billing.claim c ON c.claim_id = cl.claim_id
                    LEFT JOIN core.visit v ON v.visit_id = cl.visit_id
                    """
                ).fetchall()
                for row in rows:
                    stage = "on_track"
                    if row["status_current"] == "era_received":
                        stage = "paid"
                    expected = row["expected_amount"] or row["billed_amount"]
                    pay_date = None
                    if row["service_date"]:
                        pay_date = row["service_date"] + timedelta(days=21)
                    conn.execute(
                        """
                        INSERT INTO analytics.forecast_prediction (
                            forecast_run_id, claim_line_id, visit_id, outcome_stage,
                            expected_amount, expected_pay_date
                        )
                        VALUES (%s::uuid, %s::uuid, %s::uuid, %s, %s, %s)
                        """,
                        (
                            run_id,
                            str(row["claim_line_id"]),
                            str(row["visit_id"]) if row["visit_id"] else None,
                            stage,
                            expected,
                            pay_date,
                        ),
                    )
                    counts["predictions"] += 1
            else:
                with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as fh:
                    reader = csv.DictReader(fh)
                    for row in reader:
                        webpt_pid = safe_str(row.get("webpt_patient_id"))
                        dos = parse_date(row.get("date_of_service"))
                        cpt = safe_str(row.get("cpt_code"))
                        claim_line_id = None
                        visit_id = None
                        if webpt_pid and dos:
                            found = conn.execute(
                                """
                                SELECT cl.claim_line_id, cl.visit_id
                                FROM billing.claim_line cl
                                JOIN billing.claim c ON c.claim_id = cl.claim_id
                                JOIN core.patient p ON p.patient_id = c.patient_id
                                WHERE p.webpt_patient_id = %s
                                  AND COALESCE(cl.cpt_code, '') = COALESCE(%s, '')
                                  AND EXISTS (
                                      SELECT 1 FROM core.visit v
                                      WHERE v.visit_id = cl.visit_id AND v.service_date = %s
                                  )
                                LIMIT 1
                                """,
                                (webpt_pid, cpt, dos),
                            ).fetchone()
                            if found:
                                claim_line_id = str(found["claim_line_id"])
                                visit_id = str(found["visit_id"]) if found["visit_id"] else None
                            else:
                                v = conn.execute(
                                    """
                                    SELECT v.visit_id
                                    FROM core.visit v
                                    JOIN core.patient p ON p.patient_id = v.patient_id
                                    WHERE p.webpt_patient_id = %s AND v.service_date = %s
                                    LIMIT 1
                                    """,
                                    (webpt_pid, dos),
                                ).fetchone()
                                if v:
                                    visit_id = str(v["visit_id"])

                        stage = safe_str(row.get("outcome_stage"))
                        if stage and stage not in {
                            "paid", "on_track", "overdue", "rejected", "denied", "zero_pay"
                        }:
                            stage = "on_track"

                        risk = {
                            "source_row": {
                                k: row.get(k)
                                for k in (
                                    "risk_flag", "reconcile_status", "insurance_revflow", "source"
                                )
                                if row.get(k)
                            }
                        }
                        conn.execute(
                            """
                            INSERT INTO analytics.forecast_prediction (
                                forecast_run_id, claim_line_id, visit_id, outcome_stage,
                                expected_amount, expected_pay_date, overdue_days,
                                denied_amount, denial_category, sla_lag_days,
                                forecast_shift_days, risk_flags
                            )
                            VALUES (
                                %s::uuid, %s::uuid, %s::uuid, %s, %s, %s, %s,
                                %s, %s, %s, %s, %s::jsonb
                            )
                            """,
                            (
                                run_id,
                                claim_line_id,
                                visit_id,
                                stage,
                                parse_money(row.get("expected_amount")),
                                parse_date(row.get("expected_pay_date")),
                                int(float(row["overdue_days"])) if row.get("overdue_days") else None,
                                parse_money(row.get("denied_amount")),
                                safe_str(row.get("denial_category")),
                                int(float(row["sla_lag_days"])) if row.get("sla_lag_days") else None,
                                int(float(row["forecast_shift_days"]))
                                if row.get("forecast_shift_days")
                                else None,
                                json.dumps(risk),
                            ),
                        )
                        counts["predictions"] += 1

            finish_etl_run(conn, etl_id, status="success", row_count=counts["predictions"])
        except Exception as exc:
            finish_etl_run(conn, etl_id, status="failed", notes=str(exc)[:2000])
            raise
    return counts
