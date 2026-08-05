"""Shared Transaction Tracker xlsx parse/export helpers.

Import SoT = Month + Date columns (never sheet title).
Sheet titles are used only when exporting for workbook layout.
"""

from __future__ import annotations

import calendar
import io
import re
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, BinaryIO

from cashflow_db.util import parse_bool, parse_date, parse_money, safe_str

try:
    import openpyxl
    from openpyxl import Workbook
except ImportError:  # pragma: no cover
    openpyxl = None
    Workbook = None  # type: ignore[misc, assignment]

_WS_RE = re.compile(r"\s+")

# Logical field → accepted normalized header aliases (lowercase)
HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "payment_id": ("payment id",),
    "month_date": ("month",),
    "txn_date": ("date",),
    "amount": ("amount",),
    "eft_1": ("eft_1", "eft1"),
    "eft_2": ("eft_2", "eft2"),
    "transaction_type": ("transaction type",),
    "description": ("description",),
    "check_reference": ("#check/reference", "check/reference", "check reference"),
    "bank_name": ("bank name",),
    "billing_status": ("billing status",),
    "collector": ("collector",),
    "posted": ("posted",),
    "notes": ("notes",),
    "assigned_date": ("assigned date",),
    "claims": ("claims",),
}

REQUIRED_FIELDS = ("payment_id", "month_date", "txn_date", "amount")

CANONICAL_HEADERS = [
    "Payment ID",
    "Month",
    "Date",
    "Amount",
    "EFT_1",
    "EFT_2",
    "Transaction Type",
    "Description",
    "#Check/Reference",
    "Bank Name",
    "Billing Status",
    "Collector",
    "posted",
    "notes",
    "Assigned date",
    "CLAIMS",
]

SKIP_EFT = frozenset({"", "#N/A", "N/A", "NONE", "NULL"})


def normalize_header(cell: Any) -> str:
    if cell is None:
        return ""
    text = str(cell).replace("\xa0", " ").replace("\u200b", "")
    text = _WS_RE.sub(" ", text).strip()
    return text.lower()


def _build_col_map(header_row: tuple[Any, ...] | list[Any]) -> dict[str, int]:
    """Map logical field → column index via normalized header names."""
    norm_to_idx: dict[str, int] = {}
    for i, cell in enumerate(header_row):
        key = normalize_header(cell)
        if not key or key in norm_to_idx:
            continue
        norm_to_idx[key] = i

    col_map: dict[str, int] = {}
    for field_name, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if alias in norm_to_idx:
                col_map[field_name] = norm_to_idx[alias]
                break
    return col_map


def _cell(row: tuple[Any, ...] | list[Any], idx: int | None) -> Any:
    if idx is None or idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _clean_eft(value: Any) -> str | None:
    text = safe_str(value)
    if not text or text.upper() in SKIP_EFT:
        return None
    return text


@dataclass
class ParsedTrackerRow:
    payment_id: str
    month_date: date | None
    txn_date: date | None
    amount: Decimal | None
    eft_1: str | None = None
    eft_2: str | None = None
    transaction_type: str | None = None
    description: str | None = None
    check_reference: str | None = None
    bank_name: str | None = None
    billing_status: str | None = None
    collector: str | None = None
    posted: bool | None = None
    notes: str | None = None
    assigned_date: date | None = None
    claims: str | None = None
    source_sheet: str | None = None
    source_row: int | None = None

    def to_dict(self) -> dict[str, Any]:
        d = asdict(self)
        for k, v in list(d.items()):
            if isinstance(v, date):
                d[k] = v.isoformat()
            elif isinstance(v, Decimal):
                d[k] = str(v)
        return d


@dataclass
class SheetParseError:
    sheet: str
    message: str
    row: int | None = None


@dataclass
class ParseResult:
    rows: list[ParsedTrackerRow] = field(default_factory=list)
    errors: list[SheetParseError] = field(default_factory=list)
    skipped_sheets: list[str] = field(default_factory=list)

    @property
    def month_bounds(self) -> list[tuple[date, date]]:
        """Distinct calendar-month ranges covered by parsed Month/Date values."""
        months: set[date] = set()
        for r in self.rows:
            anchor = r.month_date or (
                r.txn_date.replace(day=1) if r.txn_date else None
            )
            if anchor:
                months.add(anchor.replace(day=1))
        bounds: list[tuple[date, date]] = []
        for start in sorted(months):
            last = calendar.monthrange(start.year, start.month)[1]
            bounds.append((start, date(start.year, start.month, last)))
        return bounds


def _row_from_cells(
    raw: tuple[Any, ...] | list[Any],
    col_map: dict[str, int],
    *,
    sheet: str,
    row_num: int,
) -> tuple[ParsedTrackerRow | None, SheetParseError | None]:
    payment_id = safe_str(_cell(raw, col_map.get("payment_id")))
    if not payment_id:
        return None, None  # blank line

    txn_date = parse_date(_cell(raw, col_map.get("txn_date")))
    month_date = parse_date(_cell(raw, col_map.get("month_date")))
    amount = parse_money(_cell(raw, col_map.get("amount")))

    if txn_date is None:
        return None, SheetParseError(
            sheet=sheet, row=row_num, message=f"Missing/invalid Date for {payment_id}"
        )
    if amount is None:
        return None, SheetParseError(
            sheet=sheet, row=row_num, message=f"Missing/invalid Amount for {payment_id}"
        )
    if month_date is None:
        month_date = txn_date.replace(day=1)

    return (
        ParsedTrackerRow(
            payment_id=payment_id,
            month_date=month_date,
            txn_date=txn_date,
            amount=amount,
            eft_1=_clean_eft(_cell(raw, col_map.get("eft_1"))),
            eft_2=_clean_eft(_cell(raw, col_map.get("eft_2"))),
            transaction_type=safe_str(_cell(raw, col_map.get("transaction_type"))),
            description=safe_str(_cell(raw, col_map.get("description"))),
            check_reference=_clean_eft(_cell(raw, col_map.get("check_reference"))),
            bank_name=safe_str(_cell(raw, col_map.get("bank_name"))),
            billing_status=safe_str(_cell(raw, col_map.get("billing_status"))),
            collector=safe_str(_cell(raw, col_map.get("collector"))),
            posted=parse_bool(_cell(raw, col_map.get("posted"))),
            notes=safe_str(_cell(raw, col_map.get("notes"))),
            assigned_date=parse_date(_cell(raw, col_map.get("assigned_date"))),
            claims=safe_str(_cell(raw, col_map.get("claims"))),
            source_sheet=sheet,
            source_row=row_num,
        ),
        None,
    )


def parse_tracker_workbook(
    source: Path | str | BinaryIO | bytes,
) -> ParseResult:
    """Parse all data sheets; skip sheets without required headers after normalize."""
    if openpyxl is None:
        raise RuntimeError("openpyxl is required to parse Transaction Tracker")

    result = ParseResult()
    if isinstance(source, (bytes, bytearray)):
        wb = openpyxl.load_workbook(io.BytesIO(source), read_only=True, data_only=True)
    elif hasattr(source, "read"):
        wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    else:
        wb = openpyxl.load_workbook(Path(source), read_only=True, data_only=True)

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            header = next(rows_iter, None)
            if header is None:
                result.skipped_sheets.append(sheet_name)
                continue
            col_map = _build_col_map(header)
            missing = [f for f in ("payment_id", "txn_date") if f not in col_map]
            if missing:
                result.skipped_sheets.append(sheet_name)
                continue
            if "amount" not in col_map:
                result.errors.append(
                    SheetParseError(
                        sheet=sheet_name,
                        message="Missing required header Amount after normalization",
                    )
                )
                continue

            for row_num, raw in enumerate(rows_iter, start=2):
                if not raw or all(c is None or str(c).strip() == "" for c in raw):
                    continue
                parsed, err = _row_from_cells(
                    raw, col_map, sheet=sheet_name, row_num=row_num
                )
                if err:
                    result.errors.append(err)
                    continue
                if parsed:
                    result.rows.append(parsed)
    finally:
        wb.close()

    # Deduplicate by payment_id (last wins) while keeping errors
    by_pid: dict[str, ParsedTrackerRow] = {}
    for r in result.rows:
        by_pid[r.payment_id] = r
    result.rows = list(by_pid.values())
    return result


def month_sheet_name(d: date) -> str:
    return calendar.month_name[d.month]


def export_tracker_workbook(rows: list[dict[str, Any]]) -> bytes:
    """Build canonical xlsx; sheet titles from txn_date for layout only."""
    if Workbook is None:
        raise RuntimeError("openpyxl is required to export Transaction Tracker")

    wb = Workbook()
    # Remove default sheet; recreate per month
    default = wb.active
    wb.remove(default)

    by_month: dict[date, list[dict[str, Any]]] = {}
    for row in rows:
        txn = row.get("txn_date")
        if isinstance(txn, str):
            txn = parse_date(txn)
        if not isinstance(txn, date):
            continue
        key = txn.replace(day=1)
        by_month.setdefault(key, []).append(row)

    if not by_month:
        ws = wb.create_sheet("Empty")
        ws.append(CANONICAL_HEADERS)
    else:
        for month_start in sorted(by_month.keys()):
            ws = wb.create_sheet(month_sheet_name(month_start))
            ws.append(CANONICAL_HEADERS)
            for row in sorted(
                by_month[month_start],
                key=lambda r: (
                    str(r.get("txn_date") or ""),
                    str(r.get("payment_id") or ""),
                ),
            ):
                month_val = row.get("month_date") or month_start
                ws.append(
                    [
                        row.get("payment_id"),
                        month_val,
                        row.get("txn_date"),
                        float(row["amount"]) if row.get("amount") is not None else None,
                        row.get("eft_1"),
                        row.get("eft_2"),
                        row.get("transaction_type"),
                        row.get("description"),
                        row.get("check_reference"),
                        row.get("bank_name"),
                        row.get("billing_status"),
                        row.get("collector"),
                        row.get("posted"),
                        row.get("notes"),
                        row.get("assigned_date"),
                        row.get("claims"),
                    ]
                )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
